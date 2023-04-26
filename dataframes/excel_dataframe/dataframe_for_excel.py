def save_excel(mapping_dict, xlsx_name, base_system_target,
               base_system_source, id_is, database):
    import pandas as pd

    tech_fields = ['changeid', 'changetype', 'changetimestamp', 'hdp_processed_dttm']
    del mapping_dict['explodedColumns']
    del mapping_dict['filter_condition']
    del mapping_dict['old_map']
    del mapping_dict['alias']
    df = pd.DataFrame(mapping_dict).sort_values(by=['table_name', 'orderBy'])
    df.drop(['orderBy'], axis=1)

    df['code_attr'] = df['code_attr'].apply(lambda x: x.replace('array', 'hash') if '_array' in x else x)

    for ind in df.index:
        code_attr = df.loc[ind, 'code_attr']
        tab_lvl = df.loc[ind, 'tab_lvl']
        colType= df.loc[ind, 'colType']
        if (colType =='hash') & ('hash' not in code_attr):
            df.loc[ind, 'code_attr'] = ''.join(code_attr.split('.')[1:])+'_hash'
        if (len(code_attr.split('_')) >= 2) & \
                (code_attr not in tech_fields) & \
                ('_hash' not in code_attr.lower()) &\
                ('_array' not in code_attr.lower()):
            df.loc[ind, 'code_attr'] = '_'.join(code_attr.split('_')[1:])
        if colType == 'hash':
            df.loc[ind, 'colType'] = 'string'

    df.insert(0, 'Схема1', f"prod_repl_subo_{database}")
    df.reset_index(inplace=True)
    df.insert(1, 'База/Система1', base_system_target)
    df['Length'] = ""
    df['PK1'] = ""
    df['FK1'] = ""
    df['Not Null1'] = ''
    df["Rejectable"] = ""
    df["Trace New Values"] = ""

    df['index'] = pd.RangeIndex(start=1, stop=len(df) + 1, step=1)

    if len(mapping_dict['short_name']) == 0:
        del mapping_dict['short_name']
        df.rename(columns={'index': '#', 'table_name': 'Таблица1', 'parent_table': 'Родительская таблица', 'code_attr': 'Код атрибута1', 'describe_attr': 'Описание атрибута1',
                           'describe_table': 'Описание таблицы1', 'comment': 'Комментарий1', 'colType': 'Тип данных1'}, inplace=True)
    else:
        df.rename(columns={'index': '#', 'table_name': 'Таблица1', 'short_name':'Сокращенное имя', 'parent_table': 'Родительская таблица',
                           'code_attr': 'Код атрибута1', 'describe_attr': 'Описание атрибута1',
                           'describe_table': 'Описание таблицы1', 'comment': 'Комментарий1', 'colType': 'Тип данных1'},
                  inplace=True)

    df.insert(1, 'Тип объекта', 'Реплика')

    def insert_row(row, number):
        if (row['Код атрибута1'] in tech_fields) or ('hash' in row['Код атрибута1']):
            return '-'
        else:
            if number == '2':
                return id_is
            elif number == '3':
                return base_system_source
            elif number == '4':
                return ''
            elif number == '5':
                return row['Таблица1']
            elif number == '6':
                return row['Описание таблицы1']
            elif number == '7':
                return row['Код атрибута1'].replace('_', '.')
            elif number == '8':
                return row['Описание атрибута1']
            elif number == '9':
                return row['Тип данных1']
            elif number == '10':
                return ''
            elif number == '11':
                return ''
            elif number == '12':
                return ''
            elif number == '13':
                return ''
            elif number == '14':
                return ''
            elif number == '15':
                return ''
            elif number == '16':
                return ''
            elif number == '17':
                return ''
            elif number == '18':
                return ''

    df.insert(2, 'ID РИС', df.apply(insert_row, axis=1, args=('2')))
    df.insert(3, 'База/Система', df.apply(insert_row, axis=1, args=('3')))
    df.insert(4, 'Схема', df.apply(insert_row, axis=1, args=('4')))
    df.insert(5, 'Таблица', df.apply(insert_row, axis=1, args=('5')))
    df.insert(6, 'Описание таблицы', df.apply(insert_row, axis=1, args=('6')))
    df.insert(7, 'Код атрибута', df.apply(insert_row, axis=1, args=('7')))
    df.insert(8, 'Краткое описание атрибута', df.apply(insert_row, axis=1, args=('8')))
    df.insert(9, 'Тип данных', df.apply(insert_row, axis=1, args=('9')))
    df.insert(10, 'Длина', df.apply(insert_row, axis=1, args=['10']))
    df.insert(11, 'PK', df.apply(insert_row, axis=1, args=['11']))
    df.insert(12, 'FK', df.apply(insert_row, axis=1, args=['12']))
    df.insert(13, 'Not Null', df.apply(insert_row, axis=1, args=['13']))
    df.insert(14, 'Dataset', df.apply(insert_row, axis=1, args=['14']))
    df.insert(15, 'Algorithm', df.apply(insert_row, axis=1, args=['15']))
    df.insert(16, 'Comment', df.apply(insert_row, axis=1, args=['16']))
    df.insert(17, 'Status', df.apply(insert_row, axis=1, args=['17']))
    df.insert(18, 'Version', df.apply(insert_row, axis=1, args=['18']))


    with pd.ExcelWriter(f'{xlsx_name}', engine='openpyxl', mode='w') as writer:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        df.to_excel(writer, sheet_name='Mapping', startrow=1, index = False)
        workbook = writer.book
        worksheet = writer.sheets['Mapping']

        border = Border(left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=Side(style='medium'))

        # Add the headers to the first row
        st_cell = worksheet.cell(row = 1, column = 1)
        st_cell.value = 'Source/Target'
        st_cell.font = Font(bold=True)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        fiolet_fill = PatternFill(start_color='B2AFE0', end_color='B2AFE0', fill_type='solid')
        for col in range(1, 3):
            worksheet.cell(row=1, column=col).fill = fiolet_fill
            worksheet.cell(row=2, column=col).fill = fiolet_fill

        source_cell = worksheet.cell(row=1, column=3)
        source_cell.value = 'Source'
        source_cell.font = Font(bold=True)
        worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=19)
        green_fill = PatternFill(start_color='84BD7F', end_color='84BD7F', fill_type='solid')
        for col in range(3, 20):
            worksheet.cell(row=1, column=col).fill = green_fill
            worksheet.cell(row=2, column=col).fill = green_fill

        target_cell = worksheet.cell(row=1, column=20)
        target_cell.value = 'Target'
        target_cell.font = Font(bold=True)
        worksheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=33)
        blue_fill = PatternFill(start_color='69CAE8', end_color='69CAE8', fill_type='solid')
        for col in range(20, df.shape[1]+1):
            worksheet.cell(row=1, column=col).fill = blue_fill
            worksheet.cell(row=2, column=col).fill = blue_fill

        vertical_cells = worksheet['O2:S2']
        for row in vertical_cells:
            for cell in row:
                cell.alignment = Alignment(textRotation=90)
        if 'Сокращенное имя' in df.columns:
            worksheet['AC2'].alignment = Alignment(textRotation=90)
        else:
            worksheet['AB2'].alignment = Alignment(textRotation=90)



        for i, col in enumerate(df.columns):
            column_length = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_length + 1

        # set widths of columns O through S to 15
        col_width = 3

        for col_num in range(15, 20):
            worksheet.column_dimensions[get_column_letter(col_num)].width = col_width
        if 'Сокращенное имя' not in df.columns:
            worksheet.column_dimensions[get_column_letter(28)].width = col_width
        else:
            worksheet.column_dimensions[get_column_letter(29)].width = col_width

        merged_cell = worksheet['A1:B1']
        for row in merged_cell:
            for cell in row:
                cell.border = border
        merged_cell = worksheet['C1:S2']
        for row in merged_cell:
            for cell in row:
                cell.border = border
        merged_cell = worksheet['T1:AI1']
        for row in merged_cell:
            for cell in row:
                cell.border = border

