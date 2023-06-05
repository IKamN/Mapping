import re
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter

def Mapping(data, base_system_source:str, database:str, file_name:str) -> None:
    # create new book excel
    wb = openpyxl.Workbook()

    # Create new execel list
    ws = wb.active
    ws.append(["start"])

    # Create headers
    source_target = ["#", "Тип объекта"]
    source = ["База/Система", "Класс", "Наименование класса","Тэг в JSON", "Описание Тэга", "Тип данных","Длина", "PK", "FK", "Not Null"]
    target = ["База/Система", "Схема", "Таблица", "Название родительской таблицы", "Описание таблицы", "Код атрибута", "Описание атрибута", "Комментарий", "Тип данных", "Length", "PK", "FK", "Not Null", "Rejectable", "Trace New Values"]
    len_sourcetarget = len(source_target)
    len_source = len(source)
    len_target = len(target)
    headers = source_target + source + target

    ws.column_dimensions['A'].width = len_sourcetarget + 1
    ws.column_dimensions['B'].width = len_source + 1
    ws.column_dimensions['C'].width = len_target + 1

    ws.append(headers)

    index = 1
    for values in data.flow_data.new_flow.tables:
        source_table = re.sub(r'^.*?_', '', values.table_name)
        schema = f"prod_repl_subo_{database}"
        for column_data in values.attributes.parsedColumns:
            tag_name = ""
            if "array" in column_data.name:
                tag_name = column_data.name.replace("array", "hash")
            elif "." not in column_data.name:
                tag_name = column_data.name
            else:
                tag_name = ".".join(column_data.name.split(".")[1:]) if column_data.colType != "hash" else ".".join(
                    column_data.name.split(".")[1:]) + "_hash"

            notnull = "+" if column_data.name.lower() in ["changeid", "changetype", "hdp_processed_dttm"] else ""
            if column_data.name.lower() in ["changeid", "changetype", "changetimestamp"]:
                comment = "Техническое поле"
                tag_json = column_data.name
                tag_descr = column_data.description
                tag_colType = ""
                attr_colType = column_data.colType
            elif column_data.name == "hdp_processed_dttm":
                comment = "Техническое поле"
                tag_json = ""
                tag_descr = ""
                tag_colType = ""
                attr_colType = column_data.colType
            elif "hash" in column_data.colType:
                comment = column_data.comment
                tag_json = ""
                tag_descr = ""
                tag_colType = ""
                attr_colType = "string"
            else:
                comment = ""
                tag_descr = f"{values.describe_table}. {column_data.description}"
                tag_colType = column_data.colType
                attr_colType = "string"
                if len(values.attributes.explodedColumns) == 1:
                    tag_json = ".".join(column_data.name.split(".")[1:])
                else:
                    tag_path = '.'.join(column_data.name.split('.')[1:])
                    arr_name = column_data.name.split('.')[0]
                    tag_json = f"{arr_name}[].{tag_path}" if len(tag_path) > 0 else f"{arr_name}[]"

            if column_data.alias is not None:
                code_attr = column_data.alias
            else:
                code_attr = column_data.name


            row_data = [index,                      # "#"
                        "Реплика",                  # "Тип объекта"
                        base_system_source,         # "База/Система"
                        data.flow_data.meta_class,  # "Класс"
                        "",                         # "Наименование класса"
                        tag_json,                   # "Тэг в JSON"
                        tag_descr,                  # "Описание Тэга"
                        tag_colType,                 # "Тип данных"
                        "",                         # "Длина"
                        "",                         # "PK"
                        "",                         # "FK"
                        "",                         # "Not Null"
                        "1642_19 Озеро данных",     # "База/Система"
                        schema,                     # "Схема"
                        values.table_name,          # "Таблица"
                        values.parent_table,        # "Название родительской таблицы"
                        values.describe_table,      # "Описание таблицы"
                        code_attr,                  # "Код атрибута"
                        column_data.description,    # "Описание атрибута"
                        comment,                    # "Комментарий"
                        attr_colType,               # "Тип данных"
                        "",                         # "Length"
                        "",                         # "PK"
                        "",                         # "FK"
                        notnull,                    # "Not Null",
                        "",                         # "Rejectable"
                        "",                         # "Trace New Values"
            ]
            index += 1
            ws.append(row_data)

    for col in ws.columns:
        max_length = 0
        column = col[1:]
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = adjusted_width

    fiolet_fill = PatternFill(start_color='B2AFE0', end_color='B2AFE0', fill_type='solid')
    orange_fill = PatternFill(start_color='ffd966', end_color='ffd966', fill_type='solid')
    blue_fill = PatternFill(start_color='A5C4E0', end_color='A5C4E0', fill_type='solid')
    green_fill = PatternFill(start_color='aed59d', end_color='aed59d', fill_type='solid')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    sourcetarget_letter = get_column_letter(len_sourcetarget)
    start_source_letter = get_column_letter(len_sourcetarget + 1)
    end_source_letter = get_column_letter(len_source+len_sourcetarget)
    start_target_letter = get_column_letter(len_sourcetarget+len_source + 1)
    end_target_letter = get_column_letter(len_sourcetarget+len_source+len_target)
    for col in range(1, len_sourcetarget + 1):
        for cell in ws[f"A{col}":f"{sourcetarget_letter}{col}"][0]:
            cell.fill = fiolet_fill
        for cell in ws[f"{start_target_letter}{col}":f"{end_target_letter}{col}"][0]:
            cell.fill = blue_fill

        for cell in ws[f"A{col}:{end_target_letter}{col}"][0]:
            cell.border = border

    for cell in ws[f"{start_source_letter}{1}":f"{end_source_letter}{1}"][0]:
        cell.fill = orange_fill
    for cell in ws[f"{start_source_letter}{2}":f"{end_source_letter}{2}"][0]:
        cell.fill = green_fill


    ws.merge_cells(f"A1:{sourcetarget_letter}1")
    ws["A1"] = "Source/Target"

    ws.merge_cells(f"{start_source_letter}1:{end_source_letter}1")
    ws[f"{start_source_letter}1"] = "Source"

    ws.merge_cells(f"{start_target_letter}1:{end_target_letter}1")
    ws[f"{start_target_letter}1"] = "Target"

    # Выравнивание текста по центру в объединенных ячейках
    center_alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].alignment = center_alignment
    ws[f'{start_source_letter}1'].alignment = center_alignment
    ws[f'{start_target_letter}1'].alignment = center_alignment

    wb.save(f"{file_name}.xlsx")